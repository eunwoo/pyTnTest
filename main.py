#-*- coding:utf-8 -*-
from serialcom import commaster
import time
import openpyxl
from openpyxl.styles.borders import Border, Side
import sys
from PySide6.QtWidgets import QApplication, QWidget, QPushButton, \
    QHBoxLayout, QVBoxLayout, QListWidget, QComboBox, QLabel, QGroupBox, QMainWindow, \
    QLineEdit, QLCDNumber, QFrame, QTabWidget, QCheckBox, QFileDialog, QDialog
from PySide6 import QtCore
from PySide6 import QtGui
from PySide6.QtGui import QFont, QIcon, QPixmap, QScreen, QClipboard, QImage
from PySide6.QtCore import Slot, Signal, QObject, QRunnable, QThreadPool, QSize, QPropertyAnimation, Property, QTimer, QEvent, QMimeData
import pyqtgraph as pg
import pyqtgraph.opengl as gl
import pyqtgraph.exporters 
import matplotlib.pyplot as plt
import json
import threading
import traceback
import numpy as np
import math
import ctypes
from collections import OrderedDict
from datetime import datetime, date
import os
import pandas as pd
from shutil import copyfile
from scipy import interpolate

# https://www.pythonguis.com/tutorials/multithreading-pyside6-applications-qthreadpool/

# 아래 2줄은 작업표시줄에 아이콪 표시되도록 하기 위함
myappid = 'mycompany.myproduct.subproduct.version' # arbitrary string
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

VERSION = "V 1.2.0"

class Visualizer(QDialog):
    def __init__(self, parent):
        super(Visualizer, self).__init__()
        self.parent = parent
        self.traces = dict()
        # self.app = QApplication(sys.argv)
        # self.guiapp = QGuiApplication(sys.argv)
        self.w = gl.GLViewWidget()
        self.w.opts['distance'] = 40
        self.w.setWindowTitle('pyqtgraph example: GLLinePlotItem')
        self.w.setGeometry(0, 110, 1200, 800)
        self.mainLayout = QHBoxLayout()
        self.mainLayout.addWidget(self.w)
        self.setLayout(self.mainLayout)
        self.setGeometry(100, 50, 1400, 800)
        # self.w.show()

        # create the background grids
        gx = gl.GLGridItem()
        gx.rotate(90, 0, 1, 0)
        gx.translate(-10, 0, 0)
        self.w.addItem(gx)
        gy = gl.GLGridItem()
        gy.rotate(90, 1, 0, 0)
        gy.translate(0, -10, 0)
        self.w.addItem(gy)
        gz = gl.GLGridItem()
        gz.translate(0, 0, -10)
        self.w.addItem(gz)

        self.n = 1000
        self.m = 1000
        self.y = np.linspace(-10, 10, self.n)
        self.x = np.linspace(-10, 10, self.m)
        self.phase = 0

        for i in range(self.n):
            yi = np.array([self.y[i]] * self.m)
            d = np.sqrt(self.x ** 2 + yi ** 2)
            # if d == 0:
            #     z = 10 * np.sin(self.phase)
            z = 10 * np.cos(d + self.phase) / (d + 1)
            pts = np.vstack([self.x, yi, z]).transpose()
            self.traces[i] = gl.GLLinePlotItem(pos=pts, color=pg.glColor(
                (i, self.n * 1.3)), width=(i + 1) / 10, antialias=True)
            self.w.addItem(self.traces[i])

    def start(self):
        return
        if (sys.flags.interactive != 1) or not hasattr(QtCore, 'PYQT_VERSION'):
            QApplication.instance().exec()
            # QtGui.QApplication.instance().exec_()

    def set_plotdata(self, name, points, color, width):
        self.traces[name].setData(pos=points, color=color, width=width)

    def update1(self):
        print('update')
        for i in range(self.n):
            yi = np.array([self.y[i]] * self.m)
            d = np.sqrt(self.x ** 2 + yi ** 2)
            z = 10 * np.cos(d + self.phase) / (d + 1)
            pts = np.vstack([self.x, yi, z]).transpose()
            self.set_plotdata(
                name=i, points=pts,
                color=pg.glColor((i, self.n * 1.3)),
                width=(i + 1) / 10
            )
            self.phase -= .003

    def animation(self):
        print('animation')
        timer = QTimer(self)
        timer.timeout.connect(self.update1)
        # timer.start(20)
        self.start()

class QHLine(QFrame):
    def __init__(self):
        super(QHLine, self).__init__()
        self.setFrameShape(QFrame.HLine)
        self.setFrameShadow(QFrame.Sunken)


class QVLine(QFrame):
    def __init__(self):
        super(QVLine, self).__init__()
        self.setFrameShape(QFrame.VLine)
        self.setFrameShadow(QFrame.Sunken)

def click(widget):  # 참고!
    class Filter(QObject):
        clicked = Signal()

        def eventFilter(self, obj, event):
            if obj == widget and event.type() == QEvent.MouseButtonPress:
                self.clicked.emit()
                return True
            return False

    filter = Filter(widget)
    widget.installEventFilter(filter)
    return filter.clicked

def InitSave():
    global line
    line = 1
def SaveTestDataNoloadOne(ws, minSpeed, maxSpeed, avgSpeed):
    global line
    line += 1
    ws[f'A{line}'] = minSpeed
    ws[f'B{line}'] = maxSpeed
    ws[f'C{line}'] = avgSpeed

def SaveTestDataOne(ws, line, Speed, Torque, Current, RPM, Voltage):
    ws[f'A{line}'] = Speed
    ws[f'B{line}'] = Torque
    ws[f'C{line}'] = Current
    ws[f'D{line}'] = abs(RPM)
    ws[f'E{line}'] = Voltage


class WorkerSignals(QObject):
    addPointCandidate = Signal(tuple)
    addPoint = Signal(tuple)
    addLine = Signal()
    finished = Signal()
    error = Signal(tuple)
    result = Signal(object)
    progress = Signal(int, int)

class StartTest(QRunnable):

    def __init__(self, fn, *args, **kwargs):
        super().__init__()
        # Store constructor arguments (re-used for processing)
        self.fn = fn
        self.args = args
        self.kwargs = kwargs
        self.signals = WorkerSignals()

        # Add the callback to our kwargs
        self.kwargs['progress_callback'] = self.signals.progress
        self.kwargs['addPoint_callback'] = self.signals.addPoint
        self.kwargs['addLine_callback'] = self.signals.addLine
        self.kwargs['addPointCandidate_callback'] = self.signals.addPointCandidate

    @Slot()
    def run(self):
        '''
        Initialise the runner function with passed args, kwargs.
        '''
        print('run')

        # Retrieve args/kwargs here; and fire processing using them
        try:
            result = self.fn(*self.args, **self.kwargs)
        except:
            traceback.print_exc()
            exctype, value = sys.exc_info()[:2]
            self.signals.error.emit((exctype, value, traceback.format_exc()))
        else:
            self.signals.result.emit(result)  # Return the result of the processing
        finally:
            self.signals.finished.emit()  # Done

class QVBoxLayout1(QVBoxLayout):
    def sizeHint(self):
        return QSize(450, 700)
    def minimumSize(self):
        return QSize(450, 700)
    def maximumSize(self):
        return QSize(500, 0)


class QLineEdit1(QLineEdit):
    def __init__(self, w = 200):
        super().__init__()
        self.w = w

    def sizeHint(self):
        return QSize(self.w, 0)
    def minimumSize(self):
        return QSize(self.w, 0)
    def maximumSize(self):
        return QSize(self.w, 0)


class QPushButton1(QPushButton):
    def __init__(self, label):
        super().__init__(label)
        self.backgroundval = QtGui.QColor(255, 255, 255)
        self.toggleOn = False
        self.anim = QPropertyAnimation(self, b"background")
        self.pressedState = False
        self.timer = QTimer(self)
        self.timer.setInterval(100)
        self.timer.timeout.connect(self.updateButtonImage)
        self.rotate = 0
        
    def updateButtonImage(self):
        # print('updateButtonImage')
        icon = QtGui.QPixmap("자동.png")
        tr = QtGui.QTransform()
        self.rotate += 10
        if self.rotate >= 360:
            self.rotate -= 360
        tr.rotate(self.rotate)
        icon = icon.transformed(tr)
        self.setIcon(icon)
        self.setIconSize(QSize(40,40))
        # half_diagonal = 50/math.sqrt(2)
        # ang_deg = self.rotate - math.floor(self.rotate/90)*90
        # print(self.rotate, ang_deg)
        # w = 2*half_diagonal * math.cos((45-ang_deg)*math.pi/180)
        # self.setIconSize(QSize(w,w))

    def sizeHint(self):
        return QSize(50, 50)
    def minimumSize(self):
        return QSize(50, 50)
    def maximumSize(self):
        return QSize(50, 50)
    def enterEvent(self, event: QtGui.QEnterEvent) -> None:
        self.anim.stop()
        self.anim.setEndValue(QtGui.QColor(100, 200, 130))
        self.anim.setDuration(500)
        self.anim.start()
        return super().enterEvent(event)
    def leaveEvent(self, event: QtCore.QEvent) -> None:
        self.anim.stop()
        self.anim.setEndValue(QtGui.QColor(255, 255, 255))
        self.anim.setDuration(500)
        self.anim.start()
        return super().leaveEvent(event)
    def readBackground(self):
        return self.backgroundval

    def pressHandler(self):
        self.timer.start()
        self.pressedState = True
        val = self.backgroundval
        red = val.red() * 1.2
        if red > 255:
            red = 255
        green = val.green() * 1.2
        if green > 255:
            green = 255
        blue = val.blue() * 1.2
        if blue > 255:
            blue = 255
        # print(red, green, blue)
        self.setStyleSheet(f'''
            QPushButton{{ border: 1px solid black; background-color: rgb({red}, {green}, {blue}); border-radius: 8px; }}
            ''')

    def releaseHandler(self):
        self.pressedState = False
        val = self.backgroundval
        red = val.red()
        green = val.green()
        blue = val.blue()
        # print(red, green, blue)
        self.setStyleSheet(f'''
            QPushButton{{ border: 1px solid black; background-color: rgb({val.red()}, {val.green()}, {val.blue()}); border-radius: 8px; }}
            ''')
        print('releaseHandler End')

    def setBackground(self, val):
        self.backgroundval = val
        red = val.red() * 1.2
        if red > 255:
            red = 255
        green = val.green() * 1.2
        if green > 255:
            green = 255
        blue = val.blue() * 1.2
        if blue > 255:
            blue = 255
        # print('setBackground')
        # print(red, green, blue)
        if self.pressedState:
            self.setStyleSheet(f'''
                QPushButton{{ 
                    border: 1px solid black; 
                    background-color: rgb({red}, {green}, {blue}); 
                    border-radius: 8px; 
                    qproperty-iconSize: 48px;
                }}
                ''')
        else:
            self.setStyleSheet(f'''
                QPushButton{{ 
                    border: 1px solid black; 
                    background-color: rgb({val.red()}, {val.green()}, {val.blue()}); 
                    border-radius: 8px;
                    qproperty-iconSize: 48px;
                }}
                QPushButton:hover{{ 
                    border: 1px solid black; 
                    background-color: rgb({val.red()}, {val.green()}, {val.blue()}); 
                    border-radius: 8px; 
                    qproperty-iconSize: 48px;
                }}
                ''')
    background = Property(QtGui.QColor, readBackground, setBackground)

    def backgroundToggle(self):
        self.toggleOn = not self.toggleOn
        val = self.backgroundval
        if self.toggleOn:
            self.setStyleSheet(f'''
                QPushButton{{ border: 1px solid black; background: white; border-radius: 8px; }}
                QPushButton:hover{{ border: 1px solid black; background-color: rgb({val.red()}, {val.green()}, {val.blue()}); border-radius: 8px; }}
                ''')
        else:
            self.setStyleSheet(f'''
                QPushButton{{ border: 1px solid black; background-color: rgb(255, 200, 130); border-radius: 8px; }}
                QPushButton:hover{{ border: 1px solid black; background-color: rgb(255, 200, 130); border-radius: 8px; }}
                ''')

class DlgSheetSelect(QDialog):    
    def __init__(self, parent = None):
        super(DlgSheetSelect, self).__init__()
        self.parent = parent
        self.setup_ui()

    def setup_ui(self):
        self.filepath = "./" + self.parent.editXlsxName.text()
        wb = openpyxl.load_workbook(self.filepath)
        print(wb.sheetnames)
        self.comboSheets = QComboBox()
        for sheet in wb.sheetnames:
            self.comboSheets.addItem(sheet)
        self.layoutTopLevel = QHBoxLayout()
        self.btnDraw = QPushButton1("그리기")
        self.layoutTopLevel.addWidget(QLabel("시트 선택"))
        self.layoutTopLevel.addWidget(self.comboSheets)
        self.layoutTopLevel.addWidget(self.btnDraw)
        self.setLayout(self.layoutTopLevel)

        self.btnDraw.clicked.connect(self.Draw)

        self.timer1 = QTimer(self)
        self.timer1.timeout.connect(self.timer1handler)
        self.timer1.start(1000)

    def timer1handler(self):
        print('timer1')

    def Draw(self):
        print('Draw')
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb[self.comboSheets.currentText()]

        data = OrderedDict()

        read_excel_row = 3
        while ws[f'A{read_excel_row}'].value:
            speed = ws[f'A{read_excel_row}'].value
            torque = ws[f'B{read_excel_row}'].value
            rpm = ws[f'D{read_excel_row}'].value
            if speed not in data:
                data[speed] = []
            data[speed].append({'torque': torque, 'rpm':rpm})
            read_excel_row += 1
        for key_speed in data:
            self.parent.AddLine()
            for value in data[key_speed]:
                torque = value['torque']
                rpm = value['rpm']
                if key_speed < 0:
                    rpm = -rpm
                self.parent.AddPoint((rpm,torque))
        self.close()
    
class Main(QWidget):

    def __init__(self):
        super(Main, self).__init__()
        self.threadpool = QThreadPool()
        self.setup_ui()
        self.timer = QTimer(self)
        self.timer.setInterval(1000)
        self.timer.timeout.connect(self.ButtonBlink)

        self.timerSpeed = QTimer(self)
        self.timerSpeed.setInterval(1000)
        self.timerSpeed.timeout.connect(self.AddPointToSpeedGraph)

        # self.timer1 = QTimer(self)
        # self.timer1.timeout.connect(self.timer1handler)
        # self.timer1.start(1000)

        self.TestPointTorqueRefbySpeed = OrderedDict()
        self.speed_t = []
        self.speed_rpm = []
        self.testPointPlannedSpeedRef = []
        self.testPointPlannedTorqueRef = []

    def timer1handler(self):
        # print('timer1')
        pass

    def closeEvent(self, event):
        global stop_thread
        # do stuff
        print('closing window')
        file_setting = open('setting.txt', 'w')
        json_data = {
            "TestInvCOM" : self.comboTestInvCOM.itemText(self.comboTestInvCOM.currentIndex()),
            "TestInvBaud" : self.comboTestInvBaud.itemText(self.comboTestInvBaud.currentIndex()),
            "TestInvID" : self.comboTestInvID.itemText(self.comboTestInvID.currentIndex()),
            "LoadInvCOM" : self.comboLoadInvCOM.itemText(self.comboLoadInvCOM.currentIndex()),
            "LoadInvBaud" : self.comboLoadInvBaud.itemText(self.comboLoadInvBaud.currentIndex()),
            "LoadInvID" : self.comboLoadInvID.itemText(self.comboLoadInvID.currentIndex()),
            "SensorCOM" : self.comboSensorCOM.itemText(self.comboSensorCOM.currentIndex()),
            "SensorBaud" : self.comboSensorBaud.itemText(self.comboSensorBaud.currentIndex()),
            "SpeedRefMax" : self.editSpeedRefMax.text(),
            "SpeedRefMin" : self.editSpeedRefMin.text(),
            "SpeedRefInc" : self.editSpeedRefInc.text(),
            "TorqueRefMax" : self.editTorqueRefMax.text(),
            "TorqueRefMin" : self.editTorqueRefMin.text(),
            "TorqueRefInc" : self.editTorqueRefInc.text(),
            "TorqueRefRatio" : self.editTorqueRefRatio.text(),
            "TorqueRefTimeZeroTorque" : self.editTorqueRefTimeZeroTorque.text(),
            "RatedCurr" : self.editRatedCurr.text(),
            "Slip" : self.editSlip.text(),
            "TestInvSpeedRefManual" : self.editTestInvSpeedRefManual.text(),
            "TestInvSpeedRefManualKpd" : self.editTestInvSpeedRefManualKpd.text(),
            "TestInvTorqueRefManual" : self.editTestInvTorqueRefManual.text(),
            "TorqueRefManual" : self.editTorqueRefManual.text(),
            "XlsxNameSaveTestResult" : self.editXlsxName.text(),
            "XlsxNameLoadTestPoint" : self.editTestPointLoadXlsxName.text(),
        }
        file_setting.write(json.dumps(json_data, indent=2))
        can_exit = True
        stop_thread = True
        if can_exit:
            self.threadpool.waitForDone(-1)
            event.accept() # let the window close
        else:
            event.ignore()

    def execute_this_fn(self, progress_callback, addPoint_callback, addLine_callback, addPointCandidate_callback):
        global TorqueMeter, LoadInverter, TestInverter, stop_thread, loadInvID, testInvID

        LoadInverter.SetStationID(int(loadInvID))
        LoadInverter.WriteRunStopSrc(b'0004')
        if self.chkTorqueModeTest.isChecked():
            LoadInverter.WriteControlMode(commaster.iV5Inverter.CONTROL_MODE_SPEED)
        else:
            LoadInverter.WriteControlMode(commaster.iV5Inverter.CONTROL_MODE_TORQUE)
        LoadInverter.WriteTorqueRef(0)

        TestInverter.SetStationID(int(testInvID))
        TestInverter.WriteRunStopSrc(b'0003')
        TestInverter.WriteFreqRefSrc(b'0006')
        if self.chkTorqueModeTest.isChecked():
            TestInverter.WriteTorqueControlMode(commaster.S100Inverter.YES)
        else:
            TestInverter.WriteTorqueControlMode(commaster.S100Inverter.NO)
        TestInverter.Run()

        InitSave()
        filepath = "./" + self.editXlsxName.text()
        try:
            wb = openpyxl.load_workbook(filepath)
        except:
            wb = openpyxl.Workbook()
            wb.save(filepath)

        excel_save_row = 2
        if self.chkTorqueModeTest.isChecked():
            try:
                ws = wb['토크모드']
                ws.delete_rows(1, ws.max_row+1)
            except:
                ws = wb.create_sheet('토크모드')
            ws[f'A{excel_save_row}'] = '시험시작'
            ws[f'B{excel_save_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
            excel_save_row += 2 #  정격토크 기입하는 줄 추가
            ws[f'A{excel_save_row}'] = 'Speed[Hz]'
            ws[f'B{excel_save_row}'] = 'Torque'
            ws[f'C{excel_save_row}'] = 'Current'
            ws[f'D{excel_save_row}'] = 'RPM'
            ws[f'E{excel_save_row}'] = 'Voltage'
            excel_save_row += 1

        elif self.chkNoloadTest.isChecked():
            try:
                ws = wb['무부하속도']
            except:
                ws = wb.create_sheet('무부하속도')
            ws[f'A{excel_save_row}'] = '시험시작'
            ws[f'B{excel_save_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
            excel_save_row += 1
            ws[f'A{excel_save_row}'] = '최소'
            ws[f'B{excel_save_row}'] = '최대'
            ws[f'C{excel_save_row}'] = '평균'
            excel_save_row += 1
            NoloadTest = OrderedDict()

        else:
            try:
                ws = wb['TN곡선']
                ws.delete_rows(1, ws.max_row+1)
            except:
                ws = wb.create_sheet('TN곡선')
            ws[f'A{excel_save_row}'] = '시험시작'
            ws[f'B{excel_save_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
            excel_save_row += 1
            ws[f'A{excel_save_row}'] = 'Speed[Hz]'
            ws[f'B{excel_save_row}'] = 'Torque'
            ws[f'C{excel_save_row}'] = 'Current'
            ws[f'D{excel_save_row}'] = 'RPM'
            ws[f'E{excel_save_row}'] = 'Voltage'
            excel_save_row += 1

        if self.chkTorqueModeTest.isChecked():
            for OneSpeedRef in self.TestPointTorqueRefbySpeed:
                TestInverter.WriteTorqueRef(self.TestPointTorqueRefbySpeed[OneSpeedRef][0])
                print('set torque ref ', self.TestPointTorqueRefbySpeed[OneSpeedRef][0])
                start_time = time.time()
                while stop_thread == False:
                    time.sleep(0.1)
                    if time.time() - start_time > 1:
                        break

                print('set speed ref ', OneSpeedRef * 60 / 2)
                if OneSpeedRef >= 0:
                    LoadInverter.Run(commaster.iV5Inverter.FWD)
                else:
                    LoadInverter.Run(commaster.iV5Inverter.REV)
                LoadInverter.WriteSpeed0(abs(OneSpeedRef) * 60 / 2)
                addLine_callback.emit()

                Hz = abs(LoadInverter.ReadSpeedHz())
                while abs(Hz - abs(OneSpeedRef)) > 0.1 and stop_thread == False:
                    Hz = abs(LoadInverter.ReadSpeedHz())
                    print(f'Output Hz={Hz:.2f}')
                    start_time = time.time()
                    while stop_thread == False:
                        time.sleep(0.1)
                        if time.time() - start_time > 1:
                            break
                if stop_thread:
                    break

                for OneTorqueRef in self.TestPointTorqueRefbySpeed[OneSpeedRef]:
                    TestInverter.WriteTorqueRef(OneTorqueRef)
                    start_time = time.time()
                    while stop_thread == False:
                        time.sleep(0.1)
                        if time.time() - start_time > 3:
                            break
                    TorqueMeterValue = TorqueMeter.Read()
                    TorqueMeasured, SpeedMeasured = TorqueMeterValue.split(',')
                    rpm = LoadInverter.ReadSpeedRpm()
                    Current = TestInverter.ReadCurrent()
                    volt = TestInverter.ReadOutputVolt()                    
                    SaveTestDataOne(ws, excel_save_row, OneSpeedRef, float(TorqueMeasured), Current, rpm, volt)
                    excel_save_row += 1
                    if OneSpeedRef >= 0:
                        addPoint_callback.emit((abs(rpm), float(TorqueMeasured)))
                    else:
                        addPoint_callback.emit((-abs(rpm), float(TorqueMeasured)))
                
                TestInverter.WriteTorqueRef(self.TestPointTorqueRefbySpeed[OneSpeedRef][0])
                print('set torque ref ', self.TestPointTorqueRefbySpeed[OneSpeedRef][0])
                start_time = time.time()
                while stop_thread == False:
                    time.sleep(0.1)
                    if time.time() - start_time > 3:
                        break

                if stop_thread:
                    break
        else:
            LoadInverter.Run(commaster.iV5Inverter.FWD)
            Interval = float(self.editTorqueRefTimeZeroTorque.text())
            for OneSpeedRef in self.TestPointTorqueRefbySpeed:
                addLine_callback.emit()
                if self.chkNoloadTest.isChecked():
                    TestInverter.WriteSpeedRef(OneSpeedRef)
                    Hz = TestInverter.ReadSpeedHz()*0.01
                    while abs(Hz - OneSpeedRef) > 1:
                        Hz = TestInverter.ReadSpeedHz()*0.01
                        print(f'Output Hz={Hz:.2f}')
                        if stop_thread:
                            break
                        time.sleep(1)
                    if stop_thread:
                        break
                    NoloadTest[OneSpeedRef] = []
                    for test in range(5):
                        time.sleep(1)
                        rpm = LoadInverter.ReadSpeedRpm()
                        NoloadTest[OneSpeedRef].append(rpm)
                        TorqueMeterValue = TorqueMeter.Read()
                        TorqueMeasured, SpeedMeasured = TorqueMeterValue.split(',')
                        addPoint_callback.emit((abs(rpm), float(TorqueMeasured)))
                        if stop_thread:
                            break
                    time.sleep(5)
                    if stop_thread:
                        break
                else:
                    LoadInverter.WriteTorqueRef(0)
                    TestInverter.WriteSpeedRef(OneSpeedRef)
                    Hz = TestInverter.ReadSpeedHz()*0.01
                    try:
                        while abs(Hz - OneSpeedRef) > 1:
                            Hz = TestInverter.ReadSpeedHz()*0.01
                            # print(f'Output Hz={Hz:.2f}')
                            if stop_thread:
                                break
                            time.sleep(1)
                        if stop_thread:
                            break
                        for OneTorqueRef in self.TestPointTorqueRefbySpeed[OneSpeedRef]:
                            
                            LoadInverter.WriteTorqueRef(OneTorqueRef)

                            time.sleep(3)
                            Current = TestInverter.ReadCurrent()
                            # print(f'Current = {Current:.1f} A')
                            TorqueMeterValue = TorqueMeter.Read()
                            TorqueMeasured, SpeedMeasured = TorqueMeterValue.split(',')
                            # print(f'Torque = {TorqueMeasured} Nm')
                            
                            rpm = LoadInverter.ReadSpeedRpm()
                            
                            # print(f'RPM = {rpm} r/min')
                            volt = TestInverter.ReadOutputVolt()
                            # print(f'Voltage = {volt} V')
                            SaveTestDataOne(ws, excel_save_row, OneSpeedRef, float(TorqueMeasured), Current, rpm, volt)
                            excel_save_row += 1
                            progress_callback.emit(50, 30)
                            addPoint_callback.emit((abs(rpm), float(TorqueMeasured)))
                            time.sleep(1)
                            if Interval != 0:
                                LoadInverter.WriteTorqueRef(0)
                                time.sleep(Interval)
                            if stop_thread:
                                break
                    except TimeoutError as e:
                        print('TimeoutError')
                        print(e)
                        break
                    except Exception as e:
                        print('exception')
                        print(e)
                        break
                    if stop_thread:
                        break

            if self.chkNoloadTest.isChecked():
                for OneSpeedRef in self.TestPointTorqueRefbySpeed:
                    print(NoloadTest[OneSpeedRef])
                    abs_v = [abs(v) for v in NoloadTest[OneSpeedRef]]
                    SaveTestDataNoloadOne(ws, min(abs_v), max(abs_v), (min(abs_v)+max(abs_v))*0.5)
        
        LoadInverter.Stop()
        TestInverter.Stop()
        
        excel_save_row = 2
        ws[f'D{excel_save_row}'] = '시험종료'
        ws[f'E{excel_save_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        wb.save(filepath)
        if self.chkTorqueModeTest.isChecked() == True:
            outfile = self.add_linearity_accuracy_eq(self.editXlsxName.text())
            outfile = self.add_motor_parameter(outfile)
            self.add_image(outfile)

        if self.chkTorqueModeTest.isChecked() == False and self.chkNoloadTest.isChecked() == False:
            self.SaveTNCurve()

        return "Done."


    def Connect(self, testInvCOM, testInvBaud, loadInvCOM, loadInvBaud, sensorCOM, sensorBaud):
        global TorqueMeter, LoadInverter, TestInverter, loadInvID, testInvID
        print('OpenPort')
        self.GetStationID()
        try:
            TorqueMeter = commaster.TorqueMeter(port = sensorCOM, baud = int(sensorBaud))
        except:
            print('torque meter connect error')
            TorqueMeter = None

        try:
            LoadInverter = commaster.iV5Inverter(port = loadInvCOM, baud = int(loadInvBaud))
            LoadInverter.SetStationID(int(loadInvID))
        except:
            print('Load Inverter connect error')
            LoadInverter = None

        try:
            TestInverter = commaster.S100Inverter(port = testInvCOM, baud = int(testInvBaud))
            TestInverter.SetStationID(int(testInvID))
            index = self.comboTestInvProtocol.currentIndex()
            print('protocol index', index)
            if index == 0:
                TestInverter.SetProtocol(commaster.Inverter.PROTOCOL_LS485)
            else:
                TestInverter.SetProtocol(commaster.Inverter.PROTOCOL_MODBUS)
        except:
            print('Test Inverter connect error')
            TestInverter = None
        self.buttonCOMOpen.setEnabled(False)

    def Disconnect(self):
        global TorqueMeter, LoadInverter, TestInverter
        if TorqueMeter:
            TorqueMeter.thread.close()
        if LoadInverter:
            LoadInverter.thread.close()
        if TestInverter:
            TestInverter.thread.close()
        self.buttonCOMOpen.setEnabled(True)

    def setup_ui(self):
        baudrates = ['9600', '19200', '38400']
        stationID = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10']


        self.setWindowTitle(" - ".join(["TN Curve 시험", VERSION]))
        self.setGeometry(100, 50, 1400, 800)
        icon = QIcon('아이콘.ico')
        self.setWindowIcon(icon)

        self.layoutTopLevel = QHBoxLayout()
        
        self.layoutRunStop = QHBoxLayout()
        self.buttonStart = QPushButton1("자 동")
        self.buttonStop = QPushButton1("정 지")
        self.layoutRunStop.addWidget(self.buttonStart)
        self.layoutRunStop.addWidget(self.buttonStop)
        font = QFont('맑은 고딕', 16)
        icon = QtGui.QPixmap("자동.png")
        tr = QtGui.QTransform()
        tr.rotate(45)
        icon = icon.transformed(tr)
        self.buttonStart.setIcon(icon)
        self.buttonStart.setIconSize(QSize(50,50))

        self.buttonStart.setStyleSheet("""
            QPushButton{ border: 1px solid black; background-color: rgb(90, 110, 220); border-radius: 8px; }
            QPushButton:hover{ border: 1px solid black; background-color: rgb(255, 200, 100); border-radius: 8px; }
        """)
        self.buttonStop.setStyleSheet("""
            QPushButton{ border: 1px solid black; background: white; border-radius: 8px; }
            QPushButton:hover{ border: 1px solid black; background-color: rgb(255, 200, 100); border-radius: 8px; }
        """)
        self.buttonStart.setFont(font)
        self.buttonStop.setFont(font)
        # 경과시간
        self.labelTimePassed = QLabel("경과시간:")

        # 속도 지령 설정
        self.layoutSpeedRef = QHBoxLayout()
        self.labelSpeedRef = QLabel("속도지령 최소:")
        self.editSpeedRefMin = QLineEdit1()
        self.labelSpeedRef1 = QLabel("최대:")
        self.editSpeedRefMax = QLineEdit1()
        self.labelSpeedRef2 = QLabel("증가:")
        self.editSpeedRefInc = QLineEdit1()
        self.layoutSpeedRef.addWidget(self.labelSpeedRef)
        self.layoutSpeedRef.addWidget(self.editSpeedRefMin)
        self.layoutSpeedRef.addWidget(self.labelSpeedRef1)
        self.layoutSpeedRef.addWidget(self.editSpeedRefMax)
        self.layoutSpeedRef.addWidget(self.labelSpeedRef2)
        self.layoutSpeedRef.addWidget(self.editSpeedRefInc)

        # 토크 지령 설정
        self.layoutTorqueRef = QHBoxLayout()
        self.labelTorqueRef = QLabel("토크지령 최소:")
        self.editTorqueRefMin = QLineEdit1()
        self.labelTorqueRef1 = QLabel("최대:")
        self.editTorqueRefMax = QLineEdit1()
        self.labelTorqueRef2 = QLabel("증가:")
        self.editTorqueRefInc = QLineEdit1()
        self.layoutTorqueRef.addWidget(self.labelTorqueRef)
        self.layoutTorqueRef.addWidget(self.editTorqueRefMin)
        self.layoutTorqueRef.addWidget(self.labelTorqueRef1)
        self.layoutTorqueRef.addWidget(self.editTorqueRefMax)
        self.layoutTorqueRef.addWidget(self.labelTorqueRef2)
        self.layoutTorqueRef.addWidget(self.editTorqueRefInc)

        # 격자 테스트 포인트 생성
        self.layoutTestPointAuto = QHBoxLayout()
        self.btnTestPointAutoGen = QPushButton("격자 테스트 포인트 생성")
        self.layoutTestPointAuto.addWidget(self.btnTestPointAutoGen, 0, QtCore.Qt.AlignmentFlag.AlignRight)

        # 테스트 포인트 불러올 파일
        self.layoutTestPointLoadXlsxName = QHBoxLayout()
        self.labelTestPointLoadXlsxName = QLabel("시험포인트 불러올 파일이름:")
        self.editTestPointLoadXlsxName = QLineEdit()
        self.layoutTestPointLoadXlsxName.addWidget(self.labelTestPointLoadXlsxName)
        self.layoutTestPointLoadXlsxName.addWidget(self.editTestPointLoadXlsxName)

        # 엑셀파일 테스트 포인트 생성 및 포인트 지우기
        self.layoutTestPointXlsxAndClear = QHBoxLayout()
        self.btnTestPointXlsx = QPushButton("엑셀 테스트 포인트 추가")
        self.btnTestPointClear = QPushButton("테스트 포인트 모두 지우기")
        self.btnTestPointScreenshot = QPushButton("스크린샷")
        self.layoutTestPointXlsxAndClear.addStretch(0)
        self.layoutTestPointXlsxAndClear.addWidget(self.btnTestPointXlsx, 0, QtCore.Qt.AlignmentFlag.AlignRight)
        self.layoutTestPointXlsxAndClear.addWidget(self.btnTestPointClear, 0, QtCore.Qt.AlignmentFlag.AlignRight)
        self.layoutTestPointXlsxAndClear.addWidget(self.btnTestPointScreenshot, 0, QtCore.Qt.AlignmentFlag.AlignRight)

        # 토크 지령 100% Nm값 설정
        self.layoutTorqueRefRatio = QHBoxLayout()
        self.labelTorqueRefRatio = QLabel("100% 토크지령 Nm환산:")
        self.editTorqueRefRatio = QLineEdit1()
        self.layoutTorqueRefRatio.addWidget(self.labelTorqueRefRatio)
        self.layoutTorqueRefRatio.addWidget(self.editTorqueRefRatio, 0, QtCore.Qt.AlignmentFlag.AlignLeft)
        self.layoutTorqueRefRatio.addStretch()

        # 토크 지령 시간 간격
        self.layoutTorqueRefTimeSetting = QHBoxLayout()
        self.labelTorqueRefTimeZeroTorque = QLabel("토크지령 영토크 시간:")
        self.editTorqueRefTimeZeroTorque = QLineEdit1()
        self.layoutTorqueRefTimeSetting.addWidget(self.labelTorqueRefTimeZeroTorque, 0, QtCore.Qt.AlignmentFlag.AlignLeft)
        self.layoutTorqueRefTimeSetting.addWidget(self.editTorqueRefTimeZeroTorque, 0, QtCore.Qt.AlignmentFlag.AlignLeft)
        self.layoutTorqueRefTimeSetting.addStretch()
        # 위 2개 가로로 배치
        self.layoutTorqueRefRatioAndTime = QHBoxLayout()
        self.layoutTorqueRefRatioAndTime.addLayout(self.layoutTorqueRefRatio)
        self.layoutTorqueRefRatioAndTime.addLayout(self.layoutTorqueRefTimeSetting)

        # 오토튜닝 결과 저장
        self.layoutSaveAutoTuneResult = QHBoxLayout()
        self.labelSlip = QLabel("슬립:")
        self.editSlip = QLineEdit1(30)
        self.labelRatedCurr = QLabel("정격전류:")
        self.editRatedCurr = QLineEdit1(30)
        self.btnSaveAutoTuneResult = QPushButton("오토튜닝 결과 저장")
        self.layoutSaveAutoTuneResult.addStretch(0)
        self.layoutSaveAutoTuneResult.addWidget(self.labelRatedCurr, 0, QtCore.Qt.AlignmentFlag.AlignRight)
        self.layoutSaveAutoTuneResult.addWidget(self.editRatedCurr, 0, QtCore.Qt.AlignmentFlag.AlignRight)
        self.layoutSaveAutoTuneResult.addWidget(self.labelSlip, 0, QtCore.Qt.AlignmentFlag.AlignRight)
        self.layoutSaveAutoTuneResult.addWidget(self.editSlip, 0, QtCore.Qt.AlignmentFlag.AlignRight)
        self.layoutSaveAutoTuneResult.addWidget(self.btnSaveAutoTuneResult, 0, QtCore.Qt.AlignmentFlag.AlignRight)

        # 속도모니터링 & 무부하 속도시험
        self.layoutSpeedMonitoringAndNoloadTest = QHBoxLayout()
        self.btnSpeedMonitoring = QPushButton("속도모니터링 시작")
        self.chkNoloadTest = QCheckBox("무부하 속도시험")
        self.chkTorqueModeTest = QCheckBox("토크모드시험")
        self.btnAnalysisTorqueModeTest = QPushButton("토크모드결과분석")
        self.layoutSpeedMonitoringAndNoloadTest.addWidget(self.btnSpeedMonitoring)
        self.layoutSpeedMonitoringAndNoloadTest.addWidget(self.chkNoloadTest)
        self.layoutSpeedMonitoringAndNoloadTest.addWidget(self.chkTorqueModeTest)
        self.layoutSpeedMonitoringAndNoloadTest.addWidget(self.btnAnalysisTorqueModeTest)

        # Test Inverter
        self.groupTestInv = QGroupBox("Test Inverter")
        self.layoutTestInv = QVBoxLayout()

        # COM port, baud container
        self.layoutTestInvInner = QHBoxLayout()
        # COM port
        self.layoutTestInvCOM = QHBoxLayout()
        self.labelTestInvCOM = QLabel("Port:")
        self.comboTestInvCOM = QComboBox()
        self.layoutTestInvCOM.addWidget(self.labelTestInvCOM)
        self.layoutTestInvCOM.addWidget(self.comboTestInvCOM)
        self.layoutTestInvInner.addLayout(self.layoutTestInvCOM)
        # ID
        self.layoutTestInvID = QHBoxLayout()
        self.labeTestInvID = QLabel("ID:")
        self.comboTestInvID = QComboBox()
        for id in stationID:
            self.comboTestInvID.addItem(id)
        self.layoutTestInvID.addWidget(self.labeTestInvID)
        self.layoutTestInvID.addWidget(self.comboTestInvID)
        self.layoutTestInvInner.addLayout(self.layoutTestInvID)
        self.layoutTestInv.addLayout(self.layoutTestInvInner)
        # 프로토콜
        self.layoutTestInvProtocol = QHBoxLayout()
        self.labelTestInvProtocol = QLabel("프로토콜:")
        self.comboTestInvProtocol = QComboBox()
        self.comboTestInvProtocol.addItem("LSBUS")
        self.comboTestInvProtocol.addItem("Modbus")
        self.layoutTestInvProtocol.addWidget(self.labelTestInvProtocol)
        self.layoutTestInvProtocol.addWidget(self.comboTestInvProtocol)
        self.layoutTestInvInner.addLayout(self.layoutTestInvProtocol)
        # baud
        self.layoutTestInvBaud = QHBoxLayout()
        self.labeTestInvBaud = QLabel("Baudrate:")
        self.comboTestInvBaud = QComboBox()
        for baudrate in baudrates:
            self.comboTestInvBaud.addItem(baudrate)
        self.layoutTestInvBaud.addWidget(self.labeTestInvBaud)
        self.layoutTestInvBaud.addWidget(self.comboTestInvBaud)
        self.layoutTestInvInner.addLayout(self.layoutTestInvBaud)
        # 수동 속도 지령
        self.layoutSpeedRefManual = QHBoxLayout()
        self.labelTestInvSpeedRefManual = QLabel("속도(통신):")
        self.editTestInvSpeedRefManual = QLineEdit1()
        self.labelTestInvSpeedRefManualKpd = QLabel("속도(키패드):")
        self.editTestInvSpeedRefManualKpd = QLineEdit1()
        self.labelTestInvTorqueRefManual = QLabel("토크:")
        self.editTestInvTorqueRefManual = QLineEdit1()
        self.layoutSpeedRefManualRight = QHBoxLayout()
        self.buttonSpeedRefManualRun = QPushButton("Run")
        self.buttonSpeedRefManualStop = QPushButton("Stop")
        self.layoutSpeedRefManualRight.addWidget(self.buttonSpeedRefManualRun, 0, QtCore.Qt.AlignmentFlag.AlignLeft)
        self.layoutSpeedRefManualRight.addWidget(self.buttonSpeedRefManualStop, 0, QtCore.Qt.AlignmentFlag.AlignLeft)
        self.layoutSpeedRefManual.addWidget(self.labelTestInvSpeedRefManual)
        self.layoutSpeedRefManual.addWidget(self.editTestInvSpeedRefManual)
        self.layoutSpeedRefManual.addWidget(self.labelTestInvSpeedRefManualKpd)
        self.layoutSpeedRefManual.addWidget(self.editTestInvSpeedRefManualKpd)
        self.layoutSpeedRefManual.addWidget(self.labelTestInvTorqueRefManual)
        self.layoutSpeedRefManual.addWidget(self.editTestInvTorqueRefManual)
        self.layoutSpeedRefManual.addLayout(self.layoutSpeedRefManualRight)
        self.layoutTestInv.addLayout(self.layoutSpeedRefManual)
        # Set Groupbox Layout
        self.groupTestInv.setLayout(self.layoutTestInv)

        # Load Inverter
        self.groupLoadInv = QGroupBox("Load Inverter")
        self.layoutLoadInv = QVBoxLayout()
        
        # COM port, baud container
        self.layoutLoadInvInner = QHBoxLayout()
        # COM port
        self.layoutLoadInvCOM = QHBoxLayout()
        self.labeLoadInvCOM = QLabel("Port:")
        self.comboLoadInvCOM = QComboBox()
        self.layoutLoadInvCOM.addWidget(self.labeLoadInvCOM)
        self.layoutLoadInvCOM.addWidget(self.comboLoadInvCOM)
        self.layoutLoadInvInner.addLayout(self.layoutLoadInvCOM)
        # baud
        self.layoutLoadInvBaud = QHBoxLayout()
        self.labeLoadInvBaud = QLabel("Baudrate:")
        self.comboLoadInvBaud = QComboBox()
        for baudrate in baudrates:
            self.comboLoadInvBaud.addItem(baudrate)
        self.layoutLoadInvBaud.addWidget(self.labeLoadInvBaud)
        self.layoutLoadInvBaud.addWidget(self.comboLoadInvBaud)
        self.layoutLoadInvInner.addLayout(self.layoutLoadInvBaud)
        # ID
        self.layoutLoadInvID = QHBoxLayout()
        self.labeLoadInvID = QLabel("ID:")
        self.comboLoadInvID = QComboBox()
        for id in stationID:
            self.comboLoadInvID.addItem(id)
        self.layoutLoadInvID.addWidget(self.labeLoadInvID)
        self.layoutLoadInvID.addWidget(self.comboLoadInvID)
        self.layoutLoadInvInner.addLayout(self.layoutLoadInvID)
        self.layoutLoadInv.addLayout(self.layoutLoadInvInner)
        # 수동 토크 지령
        self.layoutTorqueRefManual = QHBoxLayout()
        self.labelTorqueRefManual = QLabel("수동 토크 지령:")
        self.layoutTorqueRefManualRight = QHBoxLayout()
        self.editTorqueRefManual = QLineEdit1()
        self.buttonTorqueRefManualRun = QPushButton("Run")
        self.buttonTorqueRefManualStop = QPushButton("Stop")
        self.layoutTorqueRefManualRight.addWidget(self.editTorqueRefManual, 0, QtCore.Qt.AlignmentFlag.AlignLeft)
        self.layoutTorqueRefManualRight.addWidget(self.buttonTorqueRefManualRun, 0, QtCore.Qt.AlignmentFlag.AlignLeft)
        self.layoutTorqueRefManualRight.addWidget(self.buttonTorqueRefManualStop, 0, QtCore.Qt.AlignmentFlag.AlignLeft)
        self.layoutTorqueRefManual.addWidget(self.labelTorqueRefManual, 0, QtCore.Qt.AlignmentFlag.AlignLeft)
        self.layoutTorqueRefManual.addLayout(self.layoutTorqueRefManualRight)
        self.layoutLoadInv.addLayout(self.layoutTorqueRefManual)
        # Set Groupbox Layout
        self.groupLoadInv.setLayout(self.layoutLoadInv)

        # Torque Meter
        self.groupSensor = QGroupBox("Torque Meter")
        self.layoutSensor = QVBoxLayout()

        # COM port, baud container
        self.layoutSensorInner = QHBoxLayout()
        # COM port
        self.layoutSensorCOM = QHBoxLayout()
        self.labeSensorCOM = QLabel("Port:")
        self.comboSensorCOM = QComboBox()
        self.layoutSensorCOM.addWidget(self.labeSensorCOM)
        self.layoutSensorCOM.addWidget(self.comboSensorCOM)
        self.layoutSensorInner.addLayout(self.layoutSensorCOM, 1)
        # baud
        self.layoutSensorBaud = QHBoxLayout()
        self.labelSensorBaud = QLabel("Baudrate:")
        self.comboSensorBaud = QComboBox()
        for baudrate in baudrates:
            self.comboSensorBaud.addItem(baudrate)
        self.layoutSensorBaud.addWidget(self.labelSensorBaud)
        self.layoutSensorBaud.addWidget(self.comboSensorBaud)
        self.layoutSensorInner.addLayout(self.layoutSensorBaud, 1)
        self.layoutSensorInner.addStretch(1)
        self.layoutSensor.addLayout(self.layoutSensorInner)
        # Set Groupbox Layout
        self.groupSensor.setLayout(self.layoutSensor)

        # COM 포트 설정
        self.layoutPort = QHBoxLayout()
        self.buttonCOM = QPushButton("COM포트 리스트 새로고침", self)
        self.buttonCOMOpen = QPushButton("연 결", self)
        self.buttonCOMClose = QPushButton("끊 기", self)
        self.layoutPort.addWidget(self.buttonCOM)
        self.layoutPort.addWidget(self.buttonCOMOpen)
        self.layoutPort.addWidget(self.buttonCOMClose)
        self.layoutPort.setStretchFactor(self.buttonCOM, 4)
        self.layoutPort.setStretchFactor(self.buttonCOMOpen, 2)
        self.layoutPort.setStretchFactor(self.buttonCOMClose, 2)

        # COM 포트 목록
        self.listCom = QListWidget()
        self.RefreshCOMList()

        # 저장 엑셀파일 이름
        self.layoutXlsxName = QHBoxLayout()
        self.labelXlsxName = QLabel("저장할 파일이름:")
        self.editXlsxName = QLineEdit()
        self.btnXlsxName = QPushButton("열기")
        self.btnDrawResult = QPushButton("결과 그리기")
        self.layoutXlsxName.addWidget(self.labelXlsxName)
        self.layoutXlsxName.addWidget(self.editXlsxName)
        self.layoutXlsxName.addWidget(self.btnXlsxName)
        self.layoutXlsxName.addWidget(self.btnDrawResult)

        # 시트 설정
        self.layoutXlsxSheet = QHBoxLayout()
        self.labelXlsxNewSheetName = QLabel("새시트:")
        self.editXlsxNewSheetName = QLineEdit()
        self.labelXlsxExistingSheetName = QLabel("기존시트:")
        self.comboXlsxExistingSheetName = QComboBox()
        self.layoutXlsxSheet.addWidget(self.labelXlsxNewSheetName)
        self.layoutXlsxSheet.addWidget(self.editXlsxNewSheetName)
        self.layoutXlsxSheet.addWidget(self.labelXlsxExistingSheetName)
        self.layoutXlsxSheet.addWidget(self.comboXlsxExistingSheetName)


        # 이벤트 핸들러 등록
        self.btnTestPointAutoGen.clicked.connect(self.AddGridTestPoint)
        click(self.labelTestPointLoadXlsxName).connect(self.SetFileNameForLoadPoint)
        self.btnTestPointXlsx.clicked.connect(self.AddXlsxTestPoint)
        self.btnTestPointClear.clicked.connect(lambda: self.ClearLine(target='all'))
        self.btnTestPointScreenshot.clicked.connect(self.screenshot)
        self.btnSaveAutoTuneResult.clicked.connect(self.SaveAutoTune)
        self.buttonStart.clicked.connect(self.StartTestWrapper)
        self.buttonStop.clicked.connect(self.StopTest)
        self.buttonCOMOpen.clicked.connect(lambda: self.Connect(
            self.comboTestInvCOM.itemText(self.comboTestInvCOM.currentIndex()),
            self.comboTestInvBaud.itemText(self.comboTestInvBaud.currentIndex()),
            self.comboLoadInvCOM.itemText(self.comboLoadInvCOM.currentIndex()),
            self.comboLoadInvBaud.itemText(self.comboLoadInvBaud.currentIndex()),
            self.comboSensorCOM.itemText(self.comboSensorCOM.currentIndex()),
            self.comboSensorBaud.itemText(self.comboSensorBaud.currentIndex())
            ))
        self.buttonCOMClose.clicked.connect(self.Disconnect)
        self.editSlip.returnPressed.connect(self.SetSlip)
        click(self.labelRatedCurr).connect(self.GetRatedCurr)
        self.editRatedCurr.returnPressed.connect(self.SetRatedCurr)
        self.btnSpeedMonitoring.clicked.connect(self.SpeedMonitoring)
        self.btnAnalysisTorqueModeTest.clicked.connect(self.analysis_torquemode)
        # --------------------------------------------------------------
        self.buttonTorqueRefManualRun.clicked.connect(self.LoadInvManualRun)
        self.buttonTorqueRefManualStop.clicked.connect(self.LoadInvManualStop)
        self.editTorqueRefManual.returnPressed.connect(self.SetTorqueRefManual)
        click(self.labelTestInvSpeedRefManualKpd).connect(self.GetTestInvSpeedRefKpd)
        self.buttonSpeedRefManualRun.clicked.connect(self.TestInvManualRun)
        self.buttonSpeedRefManualStop.clicked.connect(self.TestInvManualStop)
        self.editTestInvSpeedRefManual.returnPressed.connect(self.SetTestInvSpeedRefManual)
        self.editTestInvSpeedRefManualKpd.returnPressed.connect(self.SetTestInvSpeedRefManualKpd)
        self.editTestInvTorqueRefManual.returnPressed.connect(self.SetTestInvTorqueRefManual)
        self.comboTestInvProtocol.currentIndexChanged.connect(self.SetTestInvProtocol)
        self.buttonCOM.clicked.connect(self.RefreshCOMList)
        # --------------------------------------------------------------
        self.btnXlsxName.clicked.connect(self.OpenXlsx)
        self.btnDrawResult.clicked.connect(self.DrawXlsxResult)

        # TN곡선과 무부하 속도 모니터링 그래프 탭 컨트롤
        self.tabGraph = QTabWidget()
        # 탭 1) TN곡선 plot window
        # https://www.pythonguis.com/tutorials/pyside6-plotting-pyqtgraph/
        self.graphWidget = pg.PlotWidget()

        # plot data: x, y values
        self.graphWidget.setBackground('w')
        styles = {'color':'b', 'font-size':'20px'}
        self.graphWidget.setLabel('left', 'Torque[Nm]', **styles)
        self.graphWidget.setLabel('bottom', 'Speed[r/min]', **styles)
        self.graphWidget.showGrid(x=True, y=True)
        self.graphWidget.setXRange(-100, 2000, padding=0)
        self.graphWidget.setYRange(-30, 30, padding=0)
        self.pen = pg.mkPen(color=(255, 0, 0), width=5, style=QtCore.Qt.PenStyle.SolidLine)
        # self.tn_data = [[[89, 88, 90, 102, 104], [-20, -10, 0, 10, 20]], [[147, 148, 150, 152, 154], [-19, -9, 0, 11, 22]]]
        # self.tn_data = [[[100, 90], [0, 10]]]
        # for one_curve in self.tn_data:
        #     self.graphWidget.plot(one_curve[0], one_curve[1], pen=self.pen, symbol='o', symbolSize=10, symbolBrush=('b'))
        self.tn_data = []
        # self.AddLine()
        # self.AddPoint((100, 0))
        # self.AddPoint((90, 10))
        # print(self.tn_data)
        self.tabGraph.addTab(self.graphWidget, QIcon('TN곡선.png'), 'TN곡선')

        # 탭 2) 속도 모니터링
        self.graphWidgetSpeed = pg.PlotWidget()
        
        self.tabGraph.addTab(self.graphWidgetSpeed, QIcon('TN곡선.png'), '실시간 속도모니터링')

        leftLayout = QVBoxLayout1()
        leftLayout.addLayout(self.layoutRunStop)
        leftLayout.addWidget(self.labelTimePassed)
        leftLayout.addLayout(self.layoutSpeedRef)
        leftLayout.addLayout(self.layoutTorqueRef)
        leftLayout.addLayout(self.layoutTestPointAuto)
        leftLayout.addLayout(self.layoutTestPointLoadXlsxName)
        leftLayout.addLayout(self.layoutTestPointXlsxAndClear)
        leftLayout.addLayout(self.layoutTorqueRefRatioAndTime)
        leftLayout.addLayout(self.layoutSaveAutoTuneResult)
        leftLayout.addLayout(self.layoutSpeedMonitoringAndNoloadTest)
        leftLayout.addWidget(QHLine())
        leftLayout.addWidget(self.groupTestInv)
        leftLayout.addWidget(self.groupLoadInv)
        leftLayout.addWidget(self.groupSensor)
        leftLayout.addLayout(self.layoutPort)
        leftLayout.addWidget(self.listCom)
        leftLayout.addLayout(self.layoutXlsxName)
        leftLayout.addLayout(self.layoutXlsxSheet)
        self.layoutTopLevel.addLayout(leftLayout)
        self.layoutTopLevel.addWidget(self.tabGraph)
        self.setLayout(self.layoutTopLevel)

        self.setStyleSheet("""
            QGroupBox {
                font: bold;
                border: 1px solid silver;
                border-radius: 6px;
                margin-top: 6px;
            }

            QGroupBox::title {
                subcontrol-origin: margin;
                left: 7px;
                padding: 0px 5px 0px 5px;
            }
            """)
        # 환경설정 백업파일 만들기
        copyfile('setting.txt', 'setting_backup.txt')
        # 환경설정 불러오기
        try:
            file_setting = open('setting.txt', 'r')
            json_string = file_setting.read()
            json_object = json.loads(json_string)
            self.comboTestInvCOM.setCurrentText(json_object['TestInvCOM'])
            self.comboTestInvBaud.setCurrentText(json_object['TestInvBaud'])
            self.comboTestInvID.setCurrentText(json_object['TestInvID'])
            self.comboLoadInvCOM.setCurrentText(json_object['LoadInvCOM'])
            self.comboLoadInvBaud.setCurrentText(json_object['LoadInvBaud'])
            self.comboLoadInvID.setCurrentText(json_object['LoadInvID'])
            self.comboSensorCOM.setCurrentText(json_object['SensorCOM'])
            self.comboSensorBaud.setCurrentText(json_object['SensorBaud'])
            self.editSpeedRefMax.setText(json_object['SpeedRefMax'])
            self.editSpeedRefMin.setText(json_object['SpeedRefMin'])
            self.editSpeedRefInc.setText(json_object['SpeedRefInc'])
            self.editTorqueRefMax.setText(json_object['TorqueRefMax'])
            self.editTorqueRefMin.setText(json_object['TorqueRefMin'])
            self.editTorqueRefInc.setText(json_object['TorqueRefInc'])
            self.editTorqueRefRatio.setText(json_object['TorqueRefRatio'])
            self.editTorqueRefTimeZeroTorque.setText(json_object['TorqueRefTimeZeroTorque'])
            self.editRatedCurr.setText(json_object['RatedCurr'])
            self.editSlip.setText(json_object['Slip'])
            self.editTestInvSpeedRefManual.setText(json_object['TestInvSpeedRefManual'])
            self.editTestInvSpeedRefManualKpd.setText(json_object['TestInvSpeedRefManualKpd'])
            self.editTestInvTorqueRefManual.setText(json_object['TestInvTorqueRefManual'])
            self.editTorqueRefManual.setText(json_object['TorqueRefManual'])
            self.editXlsxName.setText(json_object['XlsxNameSaveTestResult'])
            self.editTestPointLoadXlsxName.setText(json_object['XlsxNameLoadTestPoint'])
        except:
            pass
        self.RefreshSheetInSaveExcel()

    def RefreshSheetInSaveExcel(self):
        self.filepath = "./" + self.editXlsxName.text()
        wb = openpyxl.load_workbook(self.filepath)
        for sheet in wb.sheetnames:
            self.comboXlsxExistingSheetName.addItem(sheet)

    def SpeedMonitoring(self):
        if "시작" in self.btnSpeedMonitoring.text():
            self.speed_t = []
            self.speed_rpm = []
            self.graphWidgetSpeed.clear()
            self.timerSpeed.start()
            self.btnSpeedMonitoring.setText("속도모니터링 중지")
        else:
            self.timerSpeed.stop()
            self.btnSpeedMonitoring.setText("속도모니터링 시작")

    def TestInvManualRun(self):
        global TorqueMeter, LoadInverter, TestInverter, stop_thread, loadInvID, testInvID
        self.GetStationID()
        TestInverter.SetStationID(int(testInvID))
        TestInverter.WriteRunStopSrc(b'0003')
        TestInverter.Run()

    def TestInvManualStop(self):
        global TorqueMeter, LoadInverter, TestInverter, stop_thread, loadInvID, testInvID
        self.GetStationID()
        TestInverter.Stop()

    def LoadInvManualRun(self):
        global TorqueMeter, LoadInverter, TestInverter, stop_thread, loadInvID, testInvID
        self.GetStationID()
        print('load inv id', loadInvID)
        LoadInverter.SetStationID(int(loadInvID))
        LoadInverter.WriteRunStopSrc(b'0004')
        LoadInverter.WriteControlMode(commaster.iV5Inverter.CONTROL_MODE_TORQUE)
        LoadInverter.WriteTorqueRef(0)
        LoadInverter.Run(commaster.iV5Inverter.FWD)

    def LoadInvManualStop(self):
        global TorqueMeter, LoadInverter, TestInverter, stop_thread, loadInvID, testInvID
        self.GetStationID()
        LoadInverter.SetStationID(int(loadInvID))
        LoadInverter.Stop()

    def ButtonBlink(self):
        self.buttonStart.backgroundToggle()
        time_now = time.time()
        hours, rem = divmod(time_now - self.time_start, 3600)
        minutes, second = divmod(rem, 60)
        if hours != 0.0:
            self.labelTimePassed.setText(f'경과시간: {int(hours):2} 시간 {int(minutes):2} 분 {second:.1f} 초')
        else:
            self.labelTimePassed.setText(f'경과시간: {int(minutes):2} 분 {second:.1f} 초')

    def SetTestInvSpeedRefManual(self):
        global TestInverter
        TestInverter.WriteSpeedRef(float(self.editTestInvSpeedRefManual.text()))

    def SetTestInvSpeedRefManualKpd(self):
        global TestInverter
        TestInverter.WriteSpeedRefKpd(float(self.editTestInvSpeedRefManualKpd.text()))

    def SetTestInvTorqueRefManual(self):
        global TestInverter
        TestInverter.WriteTorqueRef(float(self.editTestInvTorqueRefManual.text()))

    def SetSlip(self):
        global TestInverter
        TestInverter.WriteSlip(float(self.editSlip.text()))
    def SetRatedCurr(self):
        global TestInverter
        TestInverter.WriteRatedCurr(float(self.editRatedCurr.text()))

    def GetRatedCurr(self):
        global TestInverter
        self.editRatedCurr.setText(f'{TestInverter.ReadRatedCurr():.1f}')

    def SetTorqueRefManual(self):
        global LoadInverter
        LoadInverter.WriteTorqueRef(float(self.editTorqueRefManual.text()))
    
    def GetTestInvSpeedRefKpd(self):
        global TestInverter
        print('GetTestInvSpeedRef')
        self.editTestInvSpeedRefManualKpd.setText(f'{TestInverter.ReadSpeedRefKpd():.2f}')
    
    def SetTestInvProtocol(self, index):
        print('SetTestInvProtocol', index)
        global TestInverter
        if 'TestInverter' in globals():
            if index == 0:
                TestInverter.SetProtocol(commaster.Inverter.PROTOCOL_LS485)
            else:
                TestInverter.SetProtocol(commaster.Inverter.PROTOCOL_MODBUS)

    def progress_fn(self, n, n1):
        print(f"{n}% done {n1}")

    def thread_complete(self):
        self.timer.stop()
        self.buttonStart.timer.stop()
        self.buttonStart.setEnabled(True)
        self.buttonStart.toggleOn = False
        self.buttonStart.backgroundToggle()
        print("THREAD COMPLETE!")

    def AddPointCandidateOnePoint(self, data, update=False):
        print('AddPointCandidateOnePoint', data)
        self.TestPointTorqueRefbySpeed[data[0]].append(data[1]) 
        print(self.TestPointTorqueRefbySpeed)
        if update:
            self.DrawPlot()

    def AddPointCandidateGrid(self, data):
        print('AddPointCandidateGrid', data)
        for one_speed in data[0]:
            for one_torque in data[1]:
                self.TestPointTorqueRefbySpeed[one_speed].append(one_torque)
        print(self.TestPointTorqueRefbySpeed)
        self.DrawPlot()

    def DrawPointCandidate(self):
        for one_speed in self.TestPointTorqueRefbySpeed:
            for one_torque in self.TestPointTorqueRefbySpeed[one_speed]:
                self.graphWidget.plot([one_speed*60/2], [one_torque*float(self.editTorqueRefRatio.text())/100], pen=None, symbol='o', symbolSize=10, symbolBrush=('g'))

    def ClearLine(self, target='all'):
        if target == 'all':
            print('ClearLine(all)')
            self.TestPointTorqueRefbySpeed = {}
            self.tn_data.clear()
            self.graphWidget.clear()
        else:
            self.tn_data.clear()
            self.graphWidget.clear()

    def AddLine(self):
        self.tn_data.append([[],[]])

    def AddPoint(self, pt):
        print('AddPoint called')
        self.tn_data[-1][0].append(pt[0])
        self.tn_data[-1][1].append(pt[1])
        self.DrawPlot()

    def DrawPlot(self):
        self.graphWidget.clear()
        self.DrawPointCandidate()
        for one_curve in self.tn_data:
            if len(one_curve[0]) > 1:
                self.graphWidget.plot(one_curve[0], one_curve[1], pen=self.pen, symbol='o', symbolSize=10, symbolBrush=('b'))
            else:
                self.graphWidget.plot(one_curve[0], one_curve[1], pen=None, symbol='o', symbolSize=10, symbolBrush=('b'))
        vb = self.graphWidget.getPlotItem().getViewBox()
        vb.enableAutoRange(enable = True)
        vb.enableAutoRange(enable = False)

    def AddPointToSpeedGraph(self):
        global LoadInverter, lock
        rpm = LoadInverter.ReadSpeedRpm()
        print(f'RPM = {rpm} r/min')
        if len(self.speed_t) == 0:
            # self.speed_t.append(int(datetime.now().strftime("%S")))
            self.speed_t.append(0)
        else:
            self.speed_t.append(self.speed_t[-1] + 1)
        self.speed_rpm.append(rpm)
        self.graphWidgetSpeed.clear()
        if len(self.speed_t) == 1:
            self.graphWidgetSpeed.plot(self.speed_t, self.speed_rpm, pen=None, symbol='o', symbolSize=5, symbolBrush=('b'))
        else:
            self.graphWidgetSpeed.plot(self.speed_t, self.speed_rpm, pen=self.pen, symbol='o', symbolSize=5, symbolBrush=('b'))

    def GetStationID(self):
        global loadInvID, testInvID
        testInvID = self.comboTestInvID.itemText(self.comboTestInvID.currentIndex())
        loadInvID = self.comboLoadInvID.itemText(self.comboLoadInvID.currentIndex())

    def StartTestWrapper(self):
        global t, stop_thread, loadInvID, testInvID
        self.GetStationID()
        stop_thread = False
        self.ClearLine(target='draw')
        self.DrawPlot()
        self.buttonStart.setEnabled(False)

        self.timer.start()
        self.buttonStart.timer.start()
        self.time_start = time.time()

        # https://www.pythonguis.com/tutorials/multithreading-pyside6-applications-qthreadpool/
        t = StartTest(self.execute_this_fn)
        t.signals.addPoint.connect(self.AddPoint)
        t.signals.addPointCandidate.connect(self.AddPointCandidateGrid)
        t.signals.addLine.connect(self.AddLine)
        t.signals.progress.connect(self.progress_fn)
        t.signals.finished.connect(self.thread_complete)

        # Execute
        self.threadpool.start(t)

    def StopTest(self):
        global t, LoadInverter, TestInverter, stop_thread
        self.buttonStop.releaseHandler()
        self.buttonStart.setEnabled(True)

        self.buttonStart.toggleOn = False
        self.buttonStart.backgroundToggle()

        LoadInverter.Stop()
        TestInverter.Stop()
        self.timer.stop()
        self.buttonStart.timer.stop()
        stop_thread = True

        if 't' in locals() and t:
            t = None

    def AddGridTestPoint(self):
        SpeedRef = np.arange(float(self.editSpeedRefMin.text()), float(self.editSpeedRefMax.text()), float(self.editSpeedRefInc.text())).tolist()
        SpeedRef.append(float(self.editSpeedRefMax.text()))
        TorqueRef = np.arange(float(self.editTorqueRefMin.text()), float(self.editTorqueRefMax.text()), float(self.editTorqueRefInc.text())).tolist()
        TorqueRef.append(float(self.editTorqueRefMax.text()))
        for one_speed in SpeedRef:
            self.TestPointTorqueRefbySpeed[one_speed] = TorqueRef
        self.DrawPlot()

    def AddXlsxTestPoint(self):
        filepath = "./" + self.editTestPointLoadXlsxName.text()
        wb = openpyxl.load_workbook(filepath)
        ws = wb['Sheet1']
        row = 2
        while True:
            if ws[f'A{row}'].value == None or ws[f'B{row}'].value == None:
                break
            print(ws[f'A{row}'].value, ws[f'B{row}'].value)
            if type(ws[f'B{row}'].value) is not str:
                self.TestPointTorqueRefbySpeed[float(ws[f'A{row}'].value)] = [ws[f'B{row}'].value]
            else:
                self.TestPointTorqueRefbySpeed[float(ws[f'A{row}'].value)] = [float(item.strip()) for item in ws[f'B{row}'].value.split(',')]
            row += 1
        print(self.TestPointTorqueRefbySpeed)
        self.DrawPlot()

    def SaveAutoTune(self):
        global TestInverter
        filepath = "./" + self.editXlsxName.text()
        try:
            wb = openpyxl.load_workbook(filepath)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            wb.save(filepath)        
            wb = openpyxl.load_workbook(filepath)
        try:
            ws = wb['오토튜닝']
            print('found sheet')
        except:
            ws = wb.create_sheet('오토튜닝')
            print('create sheet')
        ws['A1'] = 'If[A]'
        ws['B1'] = f'Rs[{TestInverter.ParaUnit("Rs")}]'
        ws['C1'] = 'Lsigma[mH]'
        ws['D1'] = 'Ls[mH]'
        ws['E1'] = 'Tr[ms]'
        print(ws.rows)
        row = len(list(ws.rows))+1
        print(row)
        ws[f'A{row}'] = TestInverter.ReadIf()
        ws[f'B{row}'] = TestInverter.ReadRs()
        ws[f'C{row}'] = TestInverter.ReadLsigma()
        ws[f'D{row}'] = TestInverter.ReadLs()
        ws[f'E{row}'] = TestInverter.ReadTr()
        wb.save(filepath)        

    def OpenXlsx(self):
        os.system(f'start excel.exe {self.editXlsxName.text()}')

    def DrawXlsxResult(self):
        print('DrawXlsxResult')
        win = DlgSheetSelect(self)
        win.setWindowTitle("시트를 선택하세요")
        win.exec()

    def SetFileNameForLoadPoint(self):
        fileName = QFileDialog.getOpenFileName(self, self.tr("Open Excel"), ".", self.tr("Excel Files (*.xlsx)"))
        if len(fileName[0].split('/')[-1]) > 0:
            self.editTestPointLoadXlsxName.setText(fileName[0].split('/')[-1])

    def SaveTNCurve(self):
        tn_curve = pd.read_excel(f'{self.editXlsxName.text()}', sheet_name='TN곡선', header=1)  # skip first row, because it's date
        print(tn_curve['Speed[Hz]'])
        print(tn_curve['Speed[Hz]'].unique())
        print(tn_curve[tn_curve['Speed[Hz]']==3])
        print(tn_curve[tn_curve['Speed[Hz]']==3].iloc[::-1])
        filepath = "./" + self.editXlsxName.text()
        try:
            wb = openpyxl.load_workbook(filepath)
        except:
            wb = openpyxl.Workbook()
            wb.save(filepath)
        try:
            ws = wb['TN곡선-형식변경']
            ws.delete_rows(1, ws.max_row+1)
        except:
            ws = wb.create_sheet('TN곡선-형식변경')
        ws.cell(row=1, column=1).value = 'Hz'
        row_no = 2
        for one_speed in tn_curve['Speed[Hz]'].unique():
            reversed = tn_curve[tn_curve['Speed[Hz]']==one_speed].iloc[::-1]
            ws.cell(row=row_no, column=1).value = one_speed
            ws.cell(row=row_no, column=2).value = '속도'
            ws.cell(row=row_no+1, column=2).value = '전류'
            col_no = 3
            for i, row in reversed.iterrows():
                ws.cell(row=row_no, column=col_no).value = row['RPM']
                ws.cell(row=row_no+1, column=col_no).value = row['Current']
                col_no += 1
            row_no += 2
        wb.save(filepath)
    
    def screenshot(self):
        print('screenshot')
        exporter = pg.exporters.ImageExporter(self.graphWidget.getPlotItem())
        # exporter.parameters()['width'] = 100
        exporter.export('filename.png')

    def add_linearity_accuracy_eq(self, file_name, sheet_name='토크모드', suffix=''):
        print('analysis_torquemode')
        wb = openpyxl.load_workbook(file_name)
        ws = wb[sheet_name]
        ws['A3'].value = '100%토크'
        ws['B3'].value = float(self.editTorqueRefRatio.text())
        ws['C3'].value = 'Nm'
        ws['F4'].value = 'Linearity[%]'
        ws['G4'].value = 'Accuracy[%]'
        ws['H4'].value = '토크지령[%]'
        process_row = 5
        while ws[f'A{process_row}'].value:
            for i, torque_ref in enumerate(self.TestPointTorqueRefbySpeed[ws[f'A{process_row}'].value]):
                ws[f'H{process_row+i}'].value = torque_ref
                ws[f'F{process_row+i}'].value = f'=B{process_row+i}/B{process_row+3}*100-ABS(H{process_row+i})'
                ws[f'F{process_row+i}'].number_format = '#,##0.00'

                ws[f'G{process_row+i}'].value = f'=(B{process_row+i}-$B$3*H{process_row+i}/100)/$B$3*100'
                ws[f'G{process_row+i}'].number_format = '#,##0.00'
            
            process_row += len(self.TestPointTorqueRefbySpeed[ws[f'A{process_row}'].value])
            # break
        filename_out = file_name.split('.')[0]+suffix+'.'+file_name.split('.')[1]
        wb.save(filename_out)
        return filename_out

    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        rows[0][0].border = Border(left=Side(style='thin'), top=Side(style='thin'))
        rows[0][-1].border = Border(right=Side(style='thin'), top=Side(style='thin'))
        for c in rows[0][1:-1]:
            c.border = Border(top=Side(style='thin'))

        for c in rows[1:-1]:
            c[0].border = Border(left=Side(style='thin'))
            c[-1].border = Border(right=Side(style='thin'))

        rows[-1][0].border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
        rows[-1][-1].border = Border(right=Side(style='thin'), bottom=Side(style='thin'))
        for c in rows[-1][1:-1]:
            c.border = Border(bottom=Side(style='thin'))

    def add_motor_parameter(self, file_name, sheet_name='토크모드', suffix=''):
        global TestInverter
        wb = openpyxl.load_workbook(file_name)
        ws = wb[sheet_name]
        ws['J2'] = 'If[A]'
        ws['K2'] = f'Rs[{TestInverter.ParaUnit("Rs")}]'
        ws['L2'] = 'Lsigma[mH]'
        ws['M2'] = 'Ls[mH]'
        ws['N2'] = 'Tr[ms]'
        ws[f'J3'] = TestInverter.ReadIf()
        ws[f'K3'] = TestInverter.ReadRs()
        ws[f'L3'] = TestInverter.ReadLsigma()
        ws[f'M3'] = TestInverter.ReadLs()
        ws[f'N3'] = TestInverter.ReadTr()
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
        self.set_border(ws, 'J2:N3')
        filename_out = file_name.split('.')[0]+suffix+'.'+file_name.split('.')[1]
        wb.save(filename_out)
        return filename_out

    def add_image(self, file_name, sheet_name='토크모드', suffix=''):
        exporter = pg.exporters.ImageExporter(self.graphWidget.getPlotItem())
        # exporter.parameters()['width'] = 100
        exporter.export('clipboard.png')
        
        wb = openpyxl.load_workbook(file_name)
        ws = wb[sheet_name]
        img = openpyxl.drawing.image.Image('clipboard.png')
        img.anchor = 'J12'        
        ws.add_image(img)
        wb.save(self.editXlsxName.text().split('.')[0]+suffix+'.'+self.editXlsxName.text().split('.')[1])

    def copy_to_clipboard(self):
        exporter = pg.exporters.ImageExporter(self.graphWidget.getPlotItem())
        # exporter.parameters()['width'] = 100
        exporter.export('clipboard.png')
        image = QImage('clipboard.png')
        clipboard = QClipboard()
        clipboard.setImage(image, QClipboard.Clipboard)

        sheet_name = '토크모드_구제어보드_역토크'
        wb = openpyxl.load_workbook(self.editXlsxName.text(), data_only=True)
        ws = wb[sheet_name]
        img = openpyxl.drawing.image.Image('clipboard.png')
        img.anchor = 'J12'        
        ws.add_image(img)
        wb.save(self.editXlsxName.text().split('.')[0]+'_test.'+self.editXlsxName.text().split('.')[1])

    def analysis_torquemode(self):
        print('analysis_torquemode')
        v = Visualizer(self)
        v.animation()
        v.exec()
        return
        wb = openpyxl.load_workbook(self.editXlsxName.text(), data_only=True)
        sheet_name = '토크모드_3차보드_역토크 (1)'
        ws = wb[sheet_name]
        data = OrderedDict()

        read_excel_row = 4
        torque_rated = float(ws[f'B2'].value)
        print(type(torque_rated))
        print(torque_rated)
        while ws[f'A{read_excel_row}'].value:
            speed = ws[f'A{read_excel_row}'].value
            torque = ws[f'B{read_excel_row}'].value
            rpm = ws[f'D{read_excel_row}'].value
            torque_ref = ws[f'H{read_excel_row}'].value
            if speed not in data:
                data[speed] = []
            data[speed].append({'torque': torque, 'rpm':rpm, 'torque_ref': torque_ref})
            read_excel_row += 1
        for key_speed in data:
            torque_100perc = data[key_speed][3]['torque']
            for value in data[key_speed]:
                torque = value['torque']
                rpm = value['rpm']
                value['linearity'] = torque / torque_100perc*100 - abs(value['torque_ref'])
                value['accuracy'] = (torque - torque_rated * value['torque_ref'] / 100)/torque_rated*100 
                if int(key_speed) < 0:
                    rpm = -rpm

        torque_test_data = pd.read_excel(f'{self.editXlsxName.text()}', sheet_name=sheet_name, header=2)  # skip first two rows
        # max of linearity
        df_linearity = torque_test_data["Linearity[%]"]
        # print(df_linearity)
        # print(np.nanmax(df_linearity.values))
        max_linearity = np.nanmax(df_linearity.values)
        # min of linearity
        min_linearity = np.nanmin(df_linearity.values)

        # max of accuracy
        df_accuracy = torque_test_data["Accuracy[%]"]
        max_accuracy = np.nanmax(df_accuracy.values)
        # min of accuracy
        min_accuracy = np.nanmin(df_accuracy.values)

        max_abs_linearity_list = []
        max_abs_accuracy_list = []
        # torque_gain_list = np.linspace(0.8, 1.2, 51).tolist()
        torque_gain_list = [0.9]
        speed_save_list = [-5, -10, -30, -40, -50, -58]
        linearity_save = []
        accuracy_save = []
        for torque_gain in torque_gain_list:
            linearity = []
            accuracy = []
            for key_speed in data:
                # print(torque_test_data.loc[torque_test_data["Speed[Hz]"]== key_speed])
            # key_speed = -55
                df = torque_test_data.loc[torque_test_data["Speed[Hz]"]== key_speed]
                # x = df['토크지령[%]'].values.tolist()
                # y = df['Torque'].values.tolist()
                # 100%까지
                x = df['토크지령[%]'].values.tolist()[0:4]
                y = df['Torque'].values.tolist()[0:4]
                f = interpolate.interp1d(x, y, fill_value='extrapolate')
                xnew = [x1*torque_gain for x1 in x]
                ynew = f(xnew)
                linearity_by_Hz = []
                accuracy_by_Hz = []
                print(key_speed)
                for i, t_ref in enumerate(x): 
                    linearity.append(ynew[i]/ynew[3]*100-abs(t_ref))
                    accuracy.append((ynew[i] - torque_rated*t_ref/100)/torque_rated*100)
                    linearity_by_Hz.append(ynew[i]/ynew[3]*100-abs(t_ref))
                    accuracy_by_Hz.append((ynew[i] - torque_rated*t_ref/100)/torque_rated*100)
                if key_speed in speed_save_list:
                    min_linearity = np.nanmin(linearity_by_Hz)
                    max_linearity = np.nanmax(linearity_by_Hz)
                    max_abs_linearity = max([abs(min_linearity), abs(max_linearity)])
                    linearity_save.append(max_abs_linearity)
                    min_accuracy = np.nanmin(accuracy_by_Hz)
                    max_accuracy = np.nanmax(accuracy_by_Hz)
                    max_abs_accuracy = max([abs(min_accuracy), abs(max_accuracy)])
                    accuracy_save.append(max_abs_accuracy)

            # print(linearity)
            min_linearity = np.nanmin(linearity)
            max_linearity = np.nanmax(linearity)
            max_abs_linearity = max([abs(min_linearity), abs(max_linearity)])
            max_abs_linearity_list.append(max_abs_linearity)
            min_accuracy = np.nanmin(accuracy)
            max_accuracy = np.nanmax(accuracy)
            max_abs_accuracy = max([abs(min_accuracy), abs(max_accuracy)])
            max_abs_accuracy_list.append(max_abs_accuracy)
        print('linearity_save')
        print([round(item, 2) for item in linearity_save])
        print('accuracy_save')
        print([round(item, 2) for item in accuracy_save])
        print(torque_gain_list, max_abs_linearity_list)
        plt.rcParams['font.family'] = 'Malgun Gothic'
        fig, ax = plt.subplots()
        ax.plot(torque_gain_list, max_abs_linearity_list, 'o', label='linearity')
        ax.plot(torque_gain_list, max_abs_accuracy_list, 'x', label='accuracy')
        ax.set(xlabel='gain[%]', ylabel='error[%]', title='S100 4kW-4 역행토크시험 linearity, accuracy')
        ax.legend()
        ax.grid()
        plt.show()



    def RefreshCOMList(self):
        self.listCom.clear()
        ports = []
        for one_info in commaster.scan():
            self.listCom.addItem(f'{one_info["PORT"]} - {one_info["DESC"]}')
            ports.append(one_info['PORT'])

        self.comboTestInvCOM.clear()            
        for port in ports:
            self.comboTestInvCOM.addItem(port)

        self.comboLoadInvCOM.clear()            
        for port in ports:
            self.comboLoadInvCOM.addItem(port)

        self.comboSensorCOM.clear()            
        for port in ports:
            self.comboSensorCOM.addItem(port)

        return ports

def main():
    global t
    app = QApplication(sys.argv)
    win = Main()
    win.show()
    app.exec()

if __name__ == '__main__':
    main()
    exit()



